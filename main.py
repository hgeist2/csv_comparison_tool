import csv
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def read_csv(file_path, key_column):
    """Reads a CSV file into a dictionary with the specified key_column as the key."""
    data = {}
    with open(file_path, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        headers = next(reader)  # Read headers
        
        # Find the key column index
        try:
            key_index = headers.index(key_column)
        except ValueError:
            raise ValueError(f"CSV file must contain a '{key_column}' column")
            
        # Create a mapping from original column positions to values
        header_to_pos = {header: i for i, header in enumerate(headers)}
            
        for row in reader:
            # Store original row values with their header names
            row_dict = {headers[i]: value for i, value in enumerate(row)}
            # Create a sorted row with sorted headers
            sorted_row = [row_dict[header] for header in sorted(headers)]
            data[row[key_index]] = sorted_row
            
    return sorted(headers), data

def compare_csv(file1, file2, key_column, ignore_columns=None):
    """
    Compares two CSV files and finds differences in matching rows.
    Returns headers along with differences for Excel report generation.
    """
    headers1, data1 = read_csv(file1, key_column)
    headers2, data2 = read_csv(file2, key_column)

    # Compare headers
    headers1_set = set(headers1)
    headers2_set = set(headers2)
    
    # Find headers that differ between files
    only_in_headers1 = headers1_set - headers2_set
    only_in_headers2 = headers2_set - headers1_set
    
    if only_in_headers1 or only_in_headers2:
        print("\nWarning: Files have different headers:")
        if only_in_headers1:
            print(f"Headers only in file 1: {sorted(only_in_headers1)}")
        if only_in_headers2:
            print(f"Headers only in file 2: {sorted(only_in_headers2)}")
        print()

    # Get indices of columns to compare (excluding ignored columns)
    ignore_columns = set(ignore_columns or [])
    compare_indices = [i for i, header in enumerate(headers1) 
                      if header not in ignore_columns]

    # Find keys unique to each file and common keys
    keys1 = set(data1.keys())
    keys2 = set(data2.keys())
    common_keys = keys1.intersection(keys2)
    only_in_file1 = keys1 - keys2
    only_in_file2 = keys2 - keys1

    differences = []
    diff_indices = []  # Track which columns have differences
    for key in common_keys:
        row1 = [data1[key][i] for i in compare_indices]
        row2 = [data2[key][i] for i in compare_indices]
        if row1 != row2:
            # Find which columns differ
            col_diffs = [i for i, (v1, v2) in enumerate(zip(row1, row2)) if v1 != v2]
            differences.append((key, data1[key], data2[key]))
            diff_indices.append(col_diffs)

    return headers1, differences, only_in_file1, only_in_file2, diff_indices

def create_excel_report(headers, differences, only_in_file1, only_in_file2, diff_indices, output_file):
    """Creates an Excel report with highlighted differences."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Differences"
    
    # Add headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add differences with highlighting
    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    current_row = 2
    
    for (key, row1, row2), diff_cols in zip(differences, diff_indices):
        # Write first file row
        for col, value in enumerate(row1, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            if col-1 in diff_cols:  # Highlight differing cells
                cell.fill = highlight_fill
        
        # Write second file row
        for col, value in enumerate(row2, 1):
            cell = ws.cell(row=current_row + 1, column=col, value=value)
            if col-1 in diff_cols:  # Highlight differing cells
                cell.fill = highlight_fill
        
        current_row += 3  # Skip a row between differences
    
    # Add missing rows information
    if only_in_file1 or only_in_file2:
        ws = wb.create_sheet("Missing Rows")
        ws.cell(row=1, column=1, value="Rows only in File 1")
        ws.cell(row=1, column=2, value="Rows only in File 2")
        
        for i, key in enumerate(sorted(only_in_file1), 2):
            ws.cell(row=i, column=1, value=key)
        for i, key in enumerate(sorted(only_in_file2), 2):
            ws.cell(row=i, column=2, value=key)
    
    wb.save(output_file)

def main():
    parser = argparse.ArgumentParser(description='Compare two CSV files')
    parser.add_argument('file1', help='Path to first CSV file')
    parser.add_argument('file2', help='Path to second CSV file')
    parser.add_argument('--key-column', default='some_id_column', 
                      help='Column to use as key for comparison (default: some_id_column)')
    parser.add_argument('--ignore', nargs='+', help='Column names to ignore in comparison')
    parser.add_argument('--output', default='comparison_results.xlsx', 
                      help='Output Excel file path (default: comparison_results.xlsx)')
    args = parser.parse_args()

    headers, differences, only_in_file1, only_in_file2, diff_indices = compare_csv(
        args.file1, 
        args.file2,
        args.key_column,
        ignore_columns=args.ignore
    )

    # Create Excel report
    create_excel_report(headers, differences, only_in_file1, only_in_file2, 
                       diff_indices, args.output)
    print(f"\nComparison results saved to {args.output}")

    # Print missing rows information
    if only_in_file1:
        print(f"\nRows only in {args.file1}: {len(only_in_file1)}")
        print("Keys:", sorted(only_in_file1))

    if only_in_file2:
        print(f"\nRows only in {args.file2}: {len(only_in_file2)}")
        print("Keys:", sorted(only_in_file2))

    # Print differences in matching rows
    if differences:
        print("\nDifferences found in matching rows:")
        for key, row1, row2 in differences:
            print(f"Key: {key}")
            print(f"File 1: {row1}")
            print(f"File 2: {row2}")
            print("-" * 40)
    else:
        print("\nNo differences found in matching rows.")

if __name__ == "__main__":
    main()
